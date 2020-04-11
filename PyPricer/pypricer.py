from yapsy.IPlugin import IPlugin  # must be imported

import io
from os import getcwd
from pathlib import Path
import json
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
import base64
import time
from functools import partial

from bokeh.models.widgets import RadioGroup, RadioButtonGroup, CheckboxGroup, Tabs, Panel, Spinner, TextInput, \
    FileInput, Select, Button, Div, MultiSelect
from bokeh.models.widgets.tables import DataTable, TableColumn
from bokeh.models.widgets.sliders import RangeSlider
from bokeh.layouts import row, column, layout
from bokeh.models import ColumnDataSource, Slider, LabelSet
from bokeh.plotting import figure

# pd.set_option('display.max_columns', None)  # or 1000
# pd.set_option('display.max_rows', None)  # or 1000
# pd.set_option('display.max_colwidth', -1)  # or 199

skp_red = '#990033'
skp_gray = '#6d6e6a'


def todict(obj, classkey=None):
    if isinstance(obj, dict):
        data = {}
        for (k, v) in obj.items():
            data[k] = todict(v, classkey)
        return data
    elif hasattr(obj, "_ast"):
        return todict(obj._ast())
    elif hasattr(obj, "__iter__") and not isinstance(obj, str):
        return [todict(v, classkey) for v in obj]
    elif hasattr(obj, "__dict__"):
        data = dict([(key, todict(value, classkey))
                     for key, value in obj.__dict__.items()
                     if not callable(value) and not key.startswith('_')])
        if classkey is not None and hasattr(obj, "__class__"):
            data[classkey] = obj.__class__.__name__
        return data
    else:
        return obj


def get_json(object):
    return json.loads(
        json.dumps(object, default=lambda o: getattr(o, '__dict__', str(o)))
    )


def get_all_tables(filename):
    """ Get all tables from a given workbook. Returns a dictionary of tables.
        Requires a filename, which includes the file path and filename.

        source: https://stackoverflow.com/a/58030207/11562752
        """

    # Load the workbook, from the filename
    wb = load_workbook(filename=filename, read_only=False, keep_vba=False, data_only=True, keep_links=False)

    # Initialize the dictionary of tables
    tables_dict = {}

    # Go through each worksheet in the workbook
    for ws_name in wb.sheetnames:
        print("")
        print(f"worksheet name: {ws_name}")
        ws = wb[ws_name]
        print(f"tables in worksheet: {len(ws._tables)}")

        # Get each table in the worksheet
        for tbl in ws._tables:
            print(f"table name: {tbl.name}")
            # First, add some info about the table to the dictionary
            tables_dict[tbl.name] = {
                'table_name': tbl.name,
                'worksheet': ws_name,
                'num_cols': len(tbl.tableColumns),
                'table_range': tbl.ref}

            # Grab the 'data' from the table
            data = ws[tbl.ref]

            # Now convert the table 'data' to a Pandas DataFrame
            # First get a list of all rows, including the first header row
            rows_list = []
            for row in data:
                # Get a list of all columns in each row
                cols = []
                for col in row:
                    cols.append(col.value)
                rows_list.append(cols)

            # Create a pandas dataframe from the rows_list.
            # The first row is the column names
            df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])

            # Add the dataframe to the dictionary of tables
            tables_dict[tbl.name]['dataframe'] = df

    return tables_dict


def get_one_table(filename, tablename):
    # Load the workbook, from the filename
    wb = load_workbook(filename=filename, read_only=False, keep_vba=False, data_only=True, keep_links=False)

    def find_tbl_data(wb):
        # Go through each worksheet in the workbook
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            for tbl in ws._tables:
                if tbl.name == tablename:
                    return ws[tbl.ref]

    data = find_tbl_data(wb)

    # Now convert the table 'data' to a Pandas DataFrame
    # First get a list of all rows, including the first header row
    rows_list = []
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)
        rows_list.append(cols)

    # Create a pandas dataframe from the rows_list. The first row is the column names
    df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])

    return df


def data_frame_from_xlsx(xlsx_file, range_name):
    """ Get a single rectangular region from the specified file.
    range_name can be a standard Excel reference ('Sheet1!A2:B7') or
    refer to a named region ('my_cells')."""
    wb = openpyxl.load_workbook(xlsx_file, data_only=True, read_only=True)
    if '!' in range_name:
        # passed a worksheet!cell reference
        ws_name, reg = range_name.split('!')
        if ws_name.startswith("'") and ws_name.endswith("'"):
            # optionally strip single quotes around sheet name
            ws_name = ws_name[1:-1]
        region = wb[ws_name][reg]
    else:
        # passed a named range; find the cells in the workbook
        full_range = wb.get_named_range(range_name)
        if full_range is None:
            raise ValueError(
                'Range "{}" not found in workbook "{}".'.format(range_name, xlsx_file)
            )
        # convert to list (openpyxl 2.3 returns a list but 2.4+ returns a generator)
        destinations = list(full_range.destinations)
        if len(destinations) > 1:
            raise ValueError(
                'Range "{}" in workbook "{}" contains more than one region.'.format(range_name, xlsx_file)
            )
        ws, reg = destinations[0]
        # convert to worksheet object (openpyxl 2.3 returns a worksheet object
        # but 2.4+ returns the name of a worksheet)
        if isinstance(ws, str):
            ws = wb[ws]
        region = ws[reg]
    df = pd.DataFrame([cell.value for cell in row] for row in region)
    df = df.rename(columns=df.iloc[0]).drop(df.index[0])

    return df


class _DataSource:
    """
        Abstract base class
    """

    def __init__(self):
        self.df = None
        self.df_col_headers = None
        self.filepath = TextInput()
        self.filepath.on_change('value', self.update_data)
        self.cols = None
        self.num_columns = Spinner(title="Number of columns", low=0, high=100, step=1, value=1, width=150)
        self.num_columns.on_change('value', self.update_num_columns)
        self.col_type_selectors = [Select(options=None)]
        self.col_selectors = [Select(options=self.df_col_headers)]

    def import_data(self):
        infile = Path(self.filepath.value)
        self.df = pd.read_csv(infile, parse_dates=True, infer_datetime_format=True, encoding='unicode_escape')
        self.df_col_headers = self.df.columns.tolist()

        print(self.df)
        print(self.df_col_headers)

    def update_data(self, attr, old, new):
        self.import_data()

    def update_num_columns(self, attr, old, new):
        for i in range(self.num_columns.value):
            if i + 1 > len(self.col_type_selectors):
                self.col_type_selectors.append(Select(options=[*self.cols]))
                self.col_selectors.append(Select(options=self.df_col_headers))
            else:
                self.col_type_selectors[i].options = [*self.cols]
                self.col_selectors[i].options = self.df_col_headers

    def update_col_options(self):
        self.col_type_selectors[0].options = [*self.cols]


class TxData(_DataSource):
    def __init__(self):
        # extend __init__ super class
        _DataSource.__init__(self)

        self.cols = \
            {
                'Invoice number': None,
                'Date': None,
                'Job number': None,
                'Item number': None,
                'Item description': None,
                'Customer id': None,
                'Customer name': None,
                'Family customer id': None,
                'Family customer name': None,
                'Gross sales': None,
                'Net sales': None,
                'Net price': None,
                'Material COGS (total)': None,
                'Material COGS (per unit)': None,
                'Labor COGS (total)': None,
                'Labor COGS (per unit)': None,
                'Var OH COGS (total)': None,
                'Var OH COGS (per unit)': None,
                'Cash discount %': None,
                'Cash discount $': None,
                'Rebate %': None,
                'Rebate $': None,
                'Freight cost (estimated)': None,
                'Freight cost (actual)': None,
                'Quantity': None,
                'Unit of measure': None,
                'Material margin %': None,
                'Material margin $': None,
                'Contribution margin %': None,
                'Contribution margin $': None
            }

        self.update_col_options()

    def clean_data(self):
        # TODO: set this up to draw from settings - may need a "re clean data" options if setting change
        self.df['Job Number'].replace([np.inf, -np.inf], np.nan, inplace=True)
        self.df['Net Price'].replace([np.inf, -np.inf], np.nan, inplace=True)
        self.df['Net Price'] = pd.to_numeric(self.df['Net Price'], errors='coerce')
        self.df = self.df.dropna(subset=['Job Number'])
        self.df['Job Number'] = self.df['Job Number'].astype(int).astype(str)
        self.df['Date'] = pd.to_datetime(self.df['Date'])
        self.df = self.df.dropna(how='any')

    def augment_data(self):
        return


class ProductMaster(_DataSource):
    def __init__(self):
        # extend __init__ super class
        _DataSource.__init__(self)

    def clean_data(self):
        return

    def augment_data(self):
        return


class CustomerMaster(_DataSource):
    def __init__(self):
        # extend __init__ super class
        _DataSource.__init__(self)

    def clean_data(self):
        return

    def augment_data(self):
        return


class Quote(_DataSource):
    def __init__(self):
        # extend __init__ super class
        _DataSource.__init__(self)

        # overwrite base class filepath with FileInput widget
        self.filepath = FileInput(accept=".xlsx")
        self.filepath.on_change('value', self.update_data)

        # add columndatasource and table
        self.cds = ColumnDataSource(data=dict(values=['']))
        self.table = DataTable(source=self.cds, columns=[TableColumn(field='', title='')], editable=True)

    def import_data(self):
        """
        This overrides the base class method

        :return:
        """
        # # decode file (received as b64 encoded) and write to xlsx locally
        decoded = base64.b64decode(self.filepath.value)

        # stream data to pandas df without saving locally
        toread = io.BytesIO()
        toread.write(decoded)  # pass your `decrypted` string as the argument here
        toread.seek(0)  # reset the pointer

        self.df = get_one_table(toread, 'quote_table').dropna(axis=0, how='all')

        cols = [TableColumn(field=_, title=_) for _ in self.df.columns]  # bokeh columns

        self.table.columns = cols
        self.cds.data = self.df

    def calc_model_prices(self, pricemodel):
        # TODO: calculate model price
        self.df['Model price'] = 1


class PriceModel:
    """
    fundamentally, this class contains a list of PriceGroup objects and related widgets and methods
    """

    def __init__(self):
        self.num_prod_groups = Spinner(title="Number of product groups", low=0, high=100, step=1, value=0, width=150)
        self.num_prod_groups.on_change('value', self.update_num_prod_groups, self.update_price_group_selector)
        self.prod_groups = [PriceGroup(name='Product group #1', pg_type='Product')]

        self.num_customer_groups = Spinner(title="Number of customer groups", low=0, high=100, step=1, value=0,
                                           width=150)
        self.num_customer_groups.on_change('value', self.update_num_cust_groups, self.update_price_group_selector)
        self.cust_groups = []

        # initialize with the number of product groups plus the number of customer groups
        self.price_groups = self.prod_groups + self.cust_groups

        self.num_price_groups = 1

        self.price_group_selector = RadioButtonGroup(labels=['Prod group #1'], active=0)

    def update_num_price_groups(self):
        self.num_price_groups = self.num_prod_groups.value + self.num_customer_groups.value

    def update_price_group_selector(self, attr, old, new):

        prod_groups = self.prod_groups[:self.num_prod_groups.value]
        cust_groups = self.cust_groups[:self.num_customer_groups.value]
        price_groups = prod_groups + cust_groups

        self.price_group_selector.labels = [_.name.value for _ in price_groups]
        self.price_group_selector.active = 0

    def update_num_prod_groups(self, attr, old, new):
        """
        on_change event for changes in the num_prod_groups spinner

        :param attr:
        :param old:
        :param new:
        :return:
        """
        self.update_num_price_groups()

        for i in range(new):
            if i + 1 > len(self.prod_groups):
                self.prod_groups.append(PriceGroup(name='Product group #{}'.format(i + 1), pg_type='Product'))
            self.prod_groups[i].name.on_change('value', self.update_price_group_selector)

        self.price_groups = self.prod_groups + self.cust_groups

    def update_num_cust_groups(self, attr, old, new):
        """
        on_change event for changes in the num_prod_groups spinner

        :param attr:
        :param old:
        :param new:
        :return:
        """
        self.update_num_price_groups()

        for i in range(new):
            if i + 1 > len(self.cust_groups):
                self.cust_groups.append(PriceGroup(name='Customer Group #{}'.format(i + 1), pg_type='Customer'))
            self.cust_groups[i].name.on_change('value', self.update_price_group_selector)
        self.price_groups = self.prod_groups + self.cust_groups


class PriceGroup:
    """
    Each price group (typically one for customers and one for each of a handful of product groups) is going to have
    several things
        - a column data source for the coefficients
        - a table associated with the cds (and table columns to go along with it)
        - a spinner to set the number of attributes
        - an attribute list, with associated file sources and column names
    """

    def __init__(self, name='', pg_type='Product'):

        self.name = TextInput(value=name, width=250)
        self.pg_type = pg_type

        self.num_attributes = Spinner(title="Number of attributes", low=1, high=100, step=1, value=1, width=150)
        self.num_attributes.on_change('value', self.update_attributes, self.update_attribute_selector)

        self.attributes = [Attribute()]
        self.attribute_selector = RadioButtonGroup(labels=['Attribute #1'], active=0)

    def update_attributes(self, attr, old, new):
        """
        on_change event for changes in the num_attributes Spinner
        """

        for i in range(new):
            if i + 1 > len(self.attributes):
                self.attributes.append(Attribute())
            self.attributes[i].name.on_change('value', self.update_attribute_selector)

    def update_attribute_selector(self, attr, old, new):
        attributes = self.attributes[:self.num_attributes.value]

        self.attribute_selector.labels = [attr.name.value for attr in attributes]
        self.attribute_selector.active = 0


class Attribute:
    def __init__(self):
        self.name = TextInput(value='Attribute name')

        self.sourcefile_options = ['', 'Tx data', 'Customer Master', 'Product master', 'Input manually']
        self.sourcefile = Select(value='', options=self.sourcefile_options)
        self.sourcefile.on_change('value', partial(self.update_sourcecolumns_options, cpq=None))
        self.sourcecolumns = Select(value='', options=[''])
        self.include_in_filters = CheckboxGroup(labels=[None], active=[0])

        self.configurations = [Configuration()]
        self.num_configs = Spinner(title="Number of configurations", low=1, high=200, step=1, value=1, width=150)
        self.num_configs.on_change('value', self.update_configs)

    def update_sourcefile(self, value='', cpq=None):
        self.sourcefile.value = value
        self.sourcefile.on_change('value', partial(self.update_sourcecolumns_options, cpq=cpq))
        self.set_sourcecolumns_options(cpq=cpq)

    def update_sourcecolumns_options(self, attr, old, new, cpq=None):
        """
        on_change event for changes in the sourcefile Select widget

        :param attr:
        :param old:
        :param new:
        :param cpq: Expects an instance of the PriceView class
        :return:
        """

        self.set_sourcecolumns_options(cpq=cpq)

    def set_sourcecolumns_options(self, cpq=None):
        col_options = None
        try:
            if self.sourcefile.value == 'Tx data':
                col_options = cpq.tx.df.columns.tolist()
            elif self.sourcefile.value == 'Customer Master':
                col_options = cpq.cust_master.df.columns.tolist()
            elif self.sourcefile.value == 'Product master':
                col_options = cpq.prod_master.df.columns.tolist()
        except AttributeError:
            pass

        self.sourcecolumns.options = col_options

    def update_configs(self, attr, old, new):
        """
        on_change event for changes in the num_configs Spinner
        """
        for i in range(new):
            if i + 1 > len(self.configurations):
                self.configurations.append(Configuration())


class Configuration:
    def __init__(self):
        self.name = TextInput(placeholder='Config name')
        self.coefficient = Spinner(value=0, step=0.0001)
        self.abs_or_pct = Select(options=['Absolute', 'Percent'], value='Absolute')


class _View:
    """
    abstract base class
    """

    def __init__(self, cpq):
        self.layout = self.create_layout(cpq)

    def create_layout(self, cpq):
        """
        dummy method designed to be overridden in child classes
        :param cpq:
        :return:
        """
        return layout(children=[row()])

    def refresh(self, cpq):
        self.layout.children = self.create_layout(cpq).children


class QuoteView(_View):

    def __init__(self, cpq):
        # extend __init__ super class
        _View.__init__(self, cpq)

    def create_layout(self, cpq):

        try:
            try:
                plot = cpq.create_plot()
            except AttributeError:
                plot = Div(text='Plot will appear here after datasources configured')

            new_layout = \
                layout(
                    children=[
                        column(
                            row(column(cpq.quote.table, width=800), column(cpq.quote.filepath), height=200),
                            row(column(*cpq.plot_filters), column(plot))
                        )
                    ]
                )
        except IndexError:
            new_layout = layout(children=[row()])
        return new_layout


class PriceMgtView(_View):
    def __init__(self, cpq):
        # extend __init__ super class
        _View.__init__(self, cpq)

    def create_layout(self, cpq):
        new_layout = layout(children=[row()])
        return new_layout


class AttrSettingsView(_View):
    def __init__(self, cpq):
        # extend __init__ super class
        _View.__init__(self, cpq)

    def create_layout(self, cpq):
        selected_pg_num = cpq.price_model.price_group_selector.active
        selected_pg = cpq.price_model.price_groups[selected_pg_num]

        prod_groups = cpq.price_model.prod_groups[:cpq.price_model.num_prod_groups.value]
        cust_groups = cpq.price_model.cust_groups[:cpq.price_model.num_customer_groups.value]

        attributes = selected_pg.attributes[:selected_pg.num_attributes.value]

        new_layout = \
            layout(
                children=[
                    column(
                        row(
                            column(
                                cpq.price_model.num_prod_groups,
                                *[pg.name for pg in prod_groups],
                                sizing_mode='fixed', width=275
                            ),
                            column(
                                cpq.price_model.num_customer_groups,
                                *[pg.name for pg in cust_groups],
                                sizing_mode='fixed', width=275
                            )
                        ),
                        row(
                            cpq.price_model.price_group_selector
                        ),
                        row(
                            selected_pg.num_attributes
                        ),
                        row(
                            column(Div(text='Attribute Name'), width=200),
                            column(Div(text='Source File'), width=200),
                            column(Div(text='Source Column'), width=200),
                            column(Div(text='Include in filters'), width=200),
                        ),
                        *[
                            row(
                                column(attribute.name, width=200),
                                column(attribute.sourcefile, width=200),
                                column(attribute.sourcecolumns, width=200),
                                column(attribute.include_in_filters, width=200)
                            )
                            for attribute in attributes
                        ],
                    )
                ]
            )

        return new_layout


class PriceModelConfigView(_View):
    def __init__(self, cpq):
        # extend __init__ super class
        _View.__init__(self, cpq)

    def create_layout(self, cpq):
        selected_prod_group_num = cpq.price_model.price_group_selector.active

        pg = cpq.price_model.price_groups[selected_prod_group_num]
        selected_attribute_num = pg.attribute_selector.active
        attr = pg.attributes[selected_attribute_num]

        new_layout = \
            layout(
                children=[
                    row(
                        column(
                            cpq.price_model.price_group_selector,
                            pg.attribute_selector,
                            attr.num_configs,
                            row(
                                column(Div(text='Configuration Name'), width=200),
                                column(Div(text='Coefficient'), width=200),
                                column(Div(text='Abs / %'), width=200)
                            ),
                            *[
                                row(
                                    column(config.name, width=200),
                                    column(config.coefficient, width=200),
                                    column(config.abs_or_pct, width=200),
                                )
                                for config in attr.configurations
                            ],
                        )
                    ),
                ]
            )
        return new_layout


class SettingsView(_View):
    def __init__(self, cpq):
        # extend __init__ super class
        _View.__init__(self, cpq)

    def create_layout(self, cpq):

        selected_datafile_num = cpq.datafile_selector.active

        selected_datafile = None
        if selected_datafile_num == 0:
            selected_datafile = cpq.tx
        elif selected_datafile_num == 1:
            selected_datafile = cpq.prod_master
        elif selected_datafile_num == 2:
            selected_datafile = cpq.cust_master

        num_columns = selected_datafile.num_columns
        col_type_selectors = selected_datafile.col_type_selectors[:num_columns.value]
        col_selectors = selected_datafile.col_selectors[:num_columns.value]

        if cpq.datafile_selector.active == 0:
            num_columns = cpq.tx.num_columns
        elif cpq.datafile_selector.active == 1:
            num_columns = cpq.prod_master.num_columns
        elif cpq.datafile_selector.active == 2:
            num_columns = cpq.cust_master.num_columns
        else:
            num_columns = Div(text='')

        new_layout = \
            layout(
                children=[
                    column(
                        row(
                            column(Div(text='Load config from file'), width=200),
                            column(cpq.config_data_input, width=300),
                            column(cpq.backup_config_button, width=200)
                        ),
                        row(
                            column(Div(text='Full path to Tx file'), width=200),
                            column(cpq.tx.filepath, width=500)
                        ),
                        row(
                            column(Div(text='Full path to product master file'), width=200),
                            column(cpq.cust_master.filepath, width=500)
                        ),
                        row(
                            column(Div(text='Full path to customer master file'), width=200),
                            column(cpq.prod_master.filepath, width=500)
                        ),
                        row(
                            cpq.datafile_selector
                        ),
                        row(
                            num_columns
                        ),
                        row(
                            column(Div(text='Select column type'), width=200),
                            column(Div(text='Select data column'), width=200)
                        ),
                        *[
                            row(
                                column(col_type_selector, width=200),
                                column(col_selector, width=200)
                            )
                            for col_type_selector, col_selector in zip(col_type_selectors, col_selectors)
                        ],
                    )
                ],
                sizing_mode='fixed',
            )
        return new_layout


class PyPricer(IPlugin):

    def __init__(self):
        # extend __init__ super class
        IPlugin.__init__(self)

        # instantiate main components
        self.tx = TxData()
        self.prod_master = ProductMaster()
        self.cust_master = CustomerMaster()
        self.quote = Quote()
        self.price_model = PriceModel()

        # cds_plot lives here rather than quote because they draw on several main components
        self.cds_plot = ColumnDataSource(data=dict(x=[0], y=[0], attr_label=[''], attr_name=['']))

        # plot_filters lives here because it's drawing from several main components
        self.plot_filters = [MultiSelect(title='Product_filter', value=[''], options=[''])]
        self.plot_filters[0].on_change('value', self.update_plot)

        # configuration backup settings
        self.backup_loc = TextInput(value=getcwd())
        self.backup_config_button = Button(label='Output config', button_type='success')
        self.backup_config_button.on_click(self.output_config)

        self.config_data_input = FileInput(accept=".json")
        self.config_data_input.on_change('value', self.load_config)

        self.datafile_selector = RadioButtonGroup(labels=['Tx data', 'Product master', 'Customer master'], active=0)
        self.datafile_selector.on_change('active', self.refresh_view)

        # define and initialize views (and their layout attributes)
        self.view_quote = QuoteView(self)
        self.view_price_mgt = PriceMgtView(self)
        self.view_price_model_config = PriceModelConfigView(self)
        self.view_attr_settings = AttrSettingsView(self)
        self.view_settings = SettingsView(self)

        self.tabs = Tabs(
            tabs=[Panel(child=self.view_quote.layout, title='Quote'),
                  Panel(child=self.view_price_mgt.layout, title='Price Management'),
                  Panel(child=self.view_price_model_config.layout, title='Price adjustments'),
                  Panel(child=self.view_attr_settings.layout, title='Attribute Settings'),
                  Panel(child=self.view_settings.layout, title='Settings')])

        # define some on_change events

        self.price_model.price_group_selector.on_change('active', self.refresh_view)
        self.price_model.num_prod_groups.on_change('value', self.refresh_view, self.update_on_change)
        self.price_model.num_customer_groups.on_change('value', self.refresh_view)
        self.tx.num_columns.on_change('value', self.refresh_view)
        self.prod_master.num_columns.on_change('value', self.refresh_view)
        self.cust_master.num_columns.on_change('value', self.refresh_view)
        for pg in self.price_model.price_groups:
            pg.num_attributes.on_change('value', self.refresh_view, self.update_on_change)
            for attr in pg.attributes:
                attr.sourcefile.on_change('value', partial(attr.update_sourcecolumns_options, cpq=self))
                attr.num_configs.on_change('value', self.refresh_view, self.update_on_change, self.update_quote_filters)


    def update_quote_filters(self, attr, old, new):
        self.set_plot_filters()

    def update_on_change(self, attr, old, new):
        for pg in self.price_model.price_groups:
            pg.num_attributes.on_change('value', self.refresh_view, self.update_on_change)
            for attr in pg.attributes:
                attr.sourcefile.on_change('value', partial(attr.update_sourcecolumns_options, cpq=self))
                attr.num_configs.on_change('value', self.refresh_view, self.update_on_change)

    def create_plot(self):
        tooltips = [
            ("Item Number", "@item_number"),
            ("Price", "@price"),
            ("CM%", "@cm_percent"),
            ("MM%", "@mm_percent"),
            ("Net Sales", "@net_sales")
        ]

        plot = figure(title='Price vs margin', plot_height=500, plot_width=500, x_range=(0, 5), y_range=[0, 1],
                      tools="save, hover, box_zoom, pan, reset, undo, redo, wheel_zoom, box_select", tooltips=tooltips)

        plot.circle('x', 'y', source=self.cds_plot, alpha=0.5)

        # clean up figure
        plot.toolbar.logo = None

        return plot

    def update_plot(self, attr, old, new):
        self.set_data_for_quote_plot()

    def output_config(self, event):

        output_dict = {}
        output_dict['tx_filepath'] = self.tx.filepath.value
        output_dict['prod_master_filepath'] = self.prod_master.filepath.value
        output_dict['cust_master_filepath'] = self.cust_master.filepath.value
        output_dict['num_prod_groups'] = self.price_model.num_prod_groups.value
        output_dict['num_customer_groups'] = self.price_model.num_customer_groups.value
        output_dict['price_groups'] = {}
        for pg in self.price_model.price_groups:
            pg_dict = {}
            pg_dict['name'] = pg.name.value
            pg_dict['num_attributes'] = pg.num_attributes.value
            pg_dict['attributes'] = {}

            for attr in pg.attributes:
                attr_dict = {}
                attr_dict['name'] = attr.name.value
                attr_dict['source_file'] = attr.sourcefile.value
                attr_dict['source_col'] = attr.sourcecolumns.value
                attr_dict['include_in_filter'] = attr.include_in_filters.active
                attr_dict['num_configs'] = attr.num_configs.value
                attr_dict['configurations'] = {}
                for config in attr.configurations:
                    config_dict = {}
                    config_dict['name'] = config.name.value
                    config_dict['coefficient'] = config.coefficient.value
                    config_dict['abs_or_pct'] = config.abs_or_pct.value

                    attr_dict['configurations'][config.name.value] = config_dict
                pg_dict['attributes'][attr.name.value] = attr_dict
            output_dict['price_groups'][pg.name.value] = pg_dict

        folder = Path(self.backup_loc.value)
        file = 'cpq_config_' + time.strftime("%Y%m%d_%H%M%S") + '.json'

        with open(folder / file, 'w') as f:
            json.dump(output_dict, f)

        print('output to', folder / file)

    def load_config(self, attr, old, new):
        # decode file (received as b64 encoded) and write to xlsx locally
        decoded = json.loads(base64.b64decode(self.config_data_input.value))

        # reset main component instances to make sure we're starting fresh
        self.tx = TxData()
        self.prod_master = ProductMaster()
        self.cust_master = CustomerMaster()
        self.quote = Quote()
        self.price_model = PriceModel()

        # data source settings
        self.tx.filepath.value = decoded['tx_filepath']
        self.prod_master.filepath.value = decoded['prod_master_filepath']
        self.cust_master.filepath.value = decoded['cust_master_filepath']

        self.price_model.num_prod_groups.value = decoded['num_prod_groups']
        self.price_model.num_customer_groups.value = decoded['num_customer_groups']
        for ((pg_name, pg_dict), pg) in zip(decoded['price_groups'].items(), self.price_model.price_groups):
            pg.name.value = pg_dict['name']
            pg.num_attributes.value = pg_dict['num_attributes']
            for ((attr_name, attr_dict), attr) in zip(pg_dict['attributes'].items(), pg.attributes):
                attr.name.value = attr_dict['name']
                attr.sourcefile.value = attr_dict['source_file']
                attr.sourcecolumns.value = attr_dict['source_col']
                attr.include_in_filters.active = attr_dict['include_in_filter']
                attr.num_configs.value = attr_dict['num_configs']
                for ((config_name, config_dict), config) in \
                        zip(attr_dict['configurations'].items(), attr.configurations):
                    config.name.value = config_dict['name']
                    config.coefficient.value = config_dict['coefficient']
                    config.abs_or_pct.value = config_dict['abs_or_pct']

        self.price_model.price_group_selector.on_change('active', self.refresh_view)
        self.price_model.num_prod_groups.on_change('value', self.refresh_view, self.update_on_change)
        self.price_model.num_customer_groups.on_change('value', self.refresh_view)
        self.tx.num_columns.on_change('value', self.refresh_view)
        self.prod_master.num_columns.on_change('value', self.refresh_view)
        self.cust_master.num_columns.on_change('value', self.refresh_view)
        for pg in self.price_model.price_groups:
            pg.num_attributes.on_change('value', self.refresh_view, self.update_on_change)
            pg.attribute_selector.on_change('active', self.refresh_view, self.update_on_change)
            for attr in pg.attributes:
                attr.sourcefile.on_change('value', partial(attr.update_sourcecolumns_options, cpq=self))
                attr.num_configs.on_change('value', self.refresh_view, self.update_on_change, self.update_quote_filters)

        self.set_plot_filters()

        self.view_quote.refresh(self)
        self.view_price_model_config.refresh(self)
        self.view_attr_settings.refresh(self)
        self.view_settings.refresh(self)

    def set_plot_filters(self):
        # TODO: Select active price group based on selected item in quote table

        selected_pg_num = 0  # self.price_model.price_group_selector.active
        selected_pg = self.price_model.price_groups[selected_pg_num]

        self.plot_filters = []
        for attr in selected_pg.attributes:
            title = attr.name.value
            values = [config.name.value for config in attr.configurations]
            self.plot_filters.append(MultiSelect(title=title, value=values, options=values))

        self.view_quote.refresh(cpq=self)

    def set_data_for_quote_plot(self):

        # For date filter, apply on Tx data

        df_item_lvl = \
            pd.pivot_table(
                self.tx.df,
                index=['Item number', 'Rollup Customer Name'],
                values=['Net Sales', 'Volume (MSI)', 'Contribution Margin $', 'Material Margin $'],
                aggfunc=np.sum
            ).reset_index()

        print('before')
        print(df_item_lvl)

        # For customer and product filters, apply on pivoted data
        # Step 1: get list of items that meet all criteria
        prod_filters = []
        for f in self.plot_filters:
            print('f.value = ', f.value)

            prod_filters.append(self.prod_master.df['Item Category Code'].isin(f.value))

        final_prod_filter = np.logical_and.reduce(prod_filters)
        included_items = self.prod_master.df.loc[final_prod_filter, 'Item Code']

        print(included_items)

        # Step 2: get list of customers that meet all criteria
        included_cust = self.cust_master.df['Parent Name']

        # Step 3; get pivoted data where items are in item list and customers are in customer list
        df_item_lvl = df_item_lvl.loc[
            df_item_lvl['Item number'].astype('str').isin(included_items.astype('str').tolist())]

        print('after')
        print(df_item_lvl)

        df_item_lvl['Net Price ($/MSI)'] = df_item_lvl['Net Sales'] / df_item_lvl['Volume (MSI)']
        df_item_lvl['CM%'] = df_item_lvl['Contribution Margin $'] / df_item_lvl['Net Sales']
        df_item_lvl['MM%'] = df_item_lvl['Material Margin $'] / df_item_lvl['Net Sales']

        df_item_lvl = df_item_lvl[
            (df_item_lvl['Net Price ($/MSI)'] > 0) &
            (df_item_lvl['Net Price ($/MSI)'] < 5) &
            (df_item_lvl['Contribution Margin $'] > -0.5) &
            (df_item_lvl['Contribution Margin $'] < 1)]

        self.cds_plot.data = \
            dict(x=df_item_lvl['Net Price ($/MSI)'],
                 y=df_item_lvl['CM%'],
                 item_number=df_item_lvl['Item number'],
                 price=df_item_lvl['Net Price ($/MSI)'],
                 cm_percent=df_item_lvl['CM%'],
                 mm_percent=df_item_lvl['Material Margin $'],
                 net_sales=df_item_lvl['Net Sales'],
                 vol_msi=df_item_lvl['Volume (MSI)'],
                 sizes=df_item_lvl['Net Sales'] / df_item_lvl['Net Sales'].max() * 0.1)

    def refresh_view(self, attr, old, new):
        self.view_price_model_config.refresh(self)
        self.view_attr_settings.refresh(self)
        self.view_settings.refresh(self)

    def bkapp(self, doc):
        doc.add_root(self.tabs)


if __name__ == '__main__':
    filepath = r'C:\Users\6008\Projects\LPX_local\lpx_tx.csv'
    infile = Path(filepath)
    df = pd.read_csv(infile, parse_dates=True, infer_datetime_format=True, encoding='unicode_escape')
    print(df.info())
